from datetime import datetime
from django.core.management.base import BaseCommand
from django.core.files import File
from openpyxl import load_workbook
import os
from home.models import *

#wb = load_workbook("myuni-data.xlsx")   

class Command(BaseCommand):
    help = 'Load data from Excel file into Django models'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='myuni-data.xlsx')

    def handle(self, *args, **options):
        file_path = options['file_path']
        wb = load_workbook(file_path)
        #sheet = wb.active 
        
        user_sheet = wb['User']
        for row in user_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == 'id' or row[0] == None:
                continue
            user = {
                    'id': int(row[0]),
                    'username': row[1],
                    'first_name': row[2],
                    'last_name': row[3],
                    'email': row[4],
                    'password': row[5],
                }
            u = User(**user)
            u.save()
            print(u)
                
                
        
        profile_sheet = wb['Profile']
        for r in profile_sheet.iter_rows(min_row=2, values_only=True):
            if r[0] == 'id' or r[0] == None:
                continue
            profile = {
                    'id': int(r[0]),
                    'user': User.objects.get(id=r[1]),
                    'profileimg': r[2],
                    'bio': r[3],
                    'location': r[4],
                }
            profile = Profile(**profile)
            profile.save()
            print(profile)
                
                
        place_sheet = wb['Place']
        for r1 in place_sheet.iter_rows(min_row=2, values_only=True):
            if r1[0] == 'id' or r1[0] == None:
                continue
            place = {
                    'id': int(r1[0]),
                    'user': User.objects.get(id=r1[1]),
                    'name': r1[2],
                    'detail': '',
                    'latitude': r1[4],
                    'longitude': r1[5],
                }
            pl = Place(**place)
            pl.save()
            print(pl)
                
                
        placeimg_sheet = wb['PlaceImage']
        for r2 in placeimg_sheet.iter_rows(min_row=2, values_only=True):
            if r2[0] == 'id' or r2[0] == None:
                continue
            place_id = r2[1]
            try:
                place = Place.objects.get(id=place_id)
                place_img = {
                    'id': int(r2[0]),
                    'place': place,
                    'image': r2[2],
                }
                place_image = PlaceImage(**place_img)
                place_image.save()
                print(place_image)    
            except Place.DoesNotExist:
                print(f"Place with ID '{place_id}' does not exist.")
                    
                
                    
        post_sheet = wb['Post']
        for r3 in post_sheet.iter_rows(min_row=2, values_only=True):
            if r3[0] == 'id' or r3[0] == None:
                    continue
            user_id = r3[1]
            place_id = r3[2]
            try:
                user = User.objects.get(id=user_id)
                place = Place.objects.get(id=place_id)
                post = {
                        'id': int(r3[0]),
                        'user': user,
                        'place': place,
                        'title': r3[3],
                        'content': r3[4],
                        'like': r3[5],       
                }
                po = Post(**post)
                po.save()
                print(po)
            except User.DoesNotExist:
                print(f"User with ID '{user_id}' does not exist.")
            except Place.DoesNotExist:
                print(f"Place with ID '{place_id}' does not exist.")
                
        
        postimg_sheet = wb['PostImage']
        for r4 in postimg_sheet.iter_rows(min_row=2, values_only=True):
            if r4[0] == 'id' or r4[0] == None:
                    continue
            post_id = r4[1]
            try:
                post = Post.objects.get(id=post_id)
                post_img = {
                    'id': int(r4[0]),
                    'post': post,
                    'image': r4[2]
                }
                post_image = PostImage(**post_img)
                post_image.save()
                print(post_image)
            except Post.DoesNotExist:
                print(f"Post with ID '{post_id}' does not exist.")
        
                
        
        postcom_sheet = wb['PostComment']
        for r5 in postcom_sheet.iter_rows(min_row=2, values_only=True):
            if r5[0] == 'id' or r5[0] == None:
                continue
            user_id = r5[1]
            post_id = r5[2]
            try:
                user = User.objects.get(id=user_id)
                post = Post.objects.get(id=post_id)
                postcom = {
                        'id': int(r5[0]),
                        'user': user,
                        'post': post,
                        'content': r5[3],
                        'like': r5[4],
                        'created_at': datetime.now(), 
                }
                postcomment = PostComment(**postcom)
                postcomment.save()
                print(postcomment)
            except User.DoesNotExist:
                print(f"User with ID '{user_id}' does not exist.")
            except Post.DoesNotExist:
                print(f"Post with ID '{post_id}' does not exist.")
                    
                
        
        postcom_img_sheet = wb['PostCommentImage']
        for r6 in postcom_img_sheet.iter_rows(min_row=2, values_only=True):
            if r6[0] == 'id' or r6[0] == None:
                continue
            post_comment_id = r6[1]
            try:
                post_comment = PostComment.objects.get(id=post_comment_id)
                postcom_img = {
                        'id': int(r6[0]),
                        'post_comment': post_comment,
                        'image': r6[2]
                }
                post_comment_image = PostCommentImage(**postcom_img)
                post_comment_image.save()
                print(post_comment_image)
            except PostComment.DoesNotExist:
                print(f"Post Comment with ID '{post_comment_id}' does not exist.")
                    
                
                    
        event_sheet = wb['Event']
        for r7 in event_sheet.iter_rows(min_row=2, values_only=True):
            if r7[0] == 'id' or r7[0] == None:
                continue
            event = {
                    'id': int(r7[0]),
                    'user': User.objects.get(id=r7[1]),
                    'place': Place.objects.get(id=r7[2]),
                    'title': r7[3],
                    'detail': r7[4],
                    'start_date': datetime.strptime(str(r7[5]), '%m/%d/%Y').strftime('%Y-%m-%d'),
                    'end_date': datetime.strptime(str(r7[6]), '%m/%d/%Y').strftime('%Y-%m-%d'),
                    'created_at': datetime.now(), 
            }
            ev = Event(**event)
            ev.save()
            print(ev)
                
                
                
        event_img_sheet = wb['EventImage']
        for r8 in event_img_sheet.iter_rows(min_row=2, values_only=True):
            if r8[0] == 'id' or r8[0] == None:
                continue
            event_id = r8[1]
            try:
                event = Event.objects.get(id=event_id)
                event_img = {
                        'id': int(r8[0]),
                        'event': event,
                        'image': r8[2],
                }
                event_image = EventImage(**event_img)
                event_image.save()
                print(event_image)
            except Event.DoesNotExist:
                print(f"Event with ID '{event_id}' does not exist.")
                